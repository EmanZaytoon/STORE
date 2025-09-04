import sqlite3
import datetime
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
import sys
import os
from openpyxl import Workbook , load_workbook

global mod_data
mod_data=[]
row_modify=1
class msg_window (QWidget):
    def __init__(self):
        super(msg_window,self).__init__()
        loadUi("msgbox.ui",self)
        self.btns()
    def btns(self):
        self.pushButton.clicked.connect(self.sd)
    def sd(self):
        #value = store_window.tableWidget.item(row, column).text()
        print(self.comboBox_9.currentText())
class modify_window (QWidget):
    def __init__(self):
        super(modify_window,self).__init__()
        loadUi("modify.ui",self)
        self.comboBox_3.setCurrentText("")
        self.comboBox_2.setCurrentText("")
        self.lineEdit_4.setText("")
        self.comboBox_4.setCurrentText("")
        self.lineEdit_5.setText(str(""))
        self.comboBox_5.setCurrentText("")
        self.lineEdit_10.setText("")
        self.lineEdit_12.setText("")
        self.connection()
        self.btns()
    def connection(self):
        table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy",
                      "الكهرباء": "electricity",
                      "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                      "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
        table = table_name[name1]
        con = sqlite3.connect("store.db")
        cursor = con.cursor()
        data = cursor.execute(f"SELECT * FROM {table} WHERE code=?", (cod,))
        for r in  data:
            ro =list(r)
            for i in range(len(ro)):
                self.comboBox_3.setCurrentText(ro[1])
                self.comboBox_2.setCurrentText(ro[4])
                self.lineEdit_4.setText(ro[2])
                self.comboBox_4.setCurrentText(ro[3])
                self.lineEdit_5.setText(str(ro[5]))
                self.comboBox_5.setCurrentText(ro[6])
                self.lineEdit_10.setText(str(ro[7]))
                self.lineEdit_12.setText(ro[9])
        con.commit()
        con.close()
    def btns(self):
        self.pushButton.clicked.connect(self.modify)
    def modify(self):
        pr_unit=0
        tota=0
        bal=0
        da=[self.lineEdit_4.text(),self.comboBox_4.currentText(),self.comboBox_2.currentText(),self.lineEdit_5.text(),self.comboBox_5.currentText(),self.lineEdit_10.text(),self.lineEdit_12.text()]
        table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy",
                      "الكهرباء": "electricity",
                      "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                      "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
        table = table_name[name1]
        con = sqlite3.connect("store.db")
        cursor = con.cursor()
        cursor.execute(f"UPDATE {table} SET item_name=? ,unit=?,state=?,amount=?,process=?,number=?,source=? WHERE code=?",(self.lineEdit_4.text(),self.comboBox_4.currentText(),self.comboBox_2.currentText(),self.lineEdit_5.text(),self.comboBox_5.currentText(),self.lineEdit_10.text(),self.lineEdit_12.text(),cod,))
        item = self.lineEdit_4.text()
        prim=cursor.execute(f"SELECT code, amount FROM msheet WHERE item_name=?",(item,))
        for code, amount in prim:
            pr_unit=float(amount)
            #for i in range (len(lr)):
        cursor.execute(f"SELECT rowid,code, amount,process FROM {table} WHERE item_name=?", (item,))
        rows = cursor.fetchall()
        for rowid ,code, amount, process in rows:
            if process == "اضافة":
                tota = tota + float(amount)
            else:
                tota = tota - float(amount)
            total = float(tota + pr_unit)
            cursor.execute(f"UPDATE {table} SET balance =? WHERE rowid=?", (total, rowid,))
        cursor.execute(f"UPDATE msheet SET balance =?  WHERE rowid= ?", (total, code,))
        self.lineEdit_7.setText(str(total))
        name = self.comboBox_3.currentText()
        wb = load_workbook("store.xlsx")
        ws2 = wb[name]
        ws3 = wb["Sheet"]
        for index, row in enumerate(ws3.iter_rows(min_row=2, values_only=True), start=2):
            val = row[2]
            if val == item:
                o = index
                ws3.cell(row=o, column=7).value = total
        m_data = [self.comboBox_3.currentText(), self.lineEdit_4.text(), self.comboBox_4.currentText(),
                  self.comboBox_2.currentText(), self.lineEdit_5.text(), self.comboBox_5.currentText(),
                  self.lineEdit_10.text(), self.dateEdit.text()
            , self.lineEdit_12.text()]
        l1 = ws2.max_column
        r1 = ws2.max_row + 1
        for i in range(2, l1):
            e = i - 2
            ws2.cell(row=row_modify, column=i).value = m_data[e]
        for i2 in range(2, r1):
            valu=ws2.cell(row=i2, column=3).value
            pross=ws2.cell(row=i2, column=7).value
            if valu == item :
                q=ws2.cell(row=i2, column=6).value
                if pross == "اضافة" :
                    bal=bal+float(q)
                else:
                    bal = bal - float(q)
                f_bal=pr_unit+bal
                ws2.cell(row=i2, column=11).value=f_bal
        wb.save("store.xlsx")
        con.commit()
        con.close()
        """m_data=[self.comboBox_3.currentText(),self.lineEdit_4.text(),self.comboBox_4.currentText(),
                self.comboBox_2.currentText() ,self.lineEdit_5.text(),self.comboBox_5.currentText(),self.lineEdit_10.text(),self.dateEdit.text()
            ,self.lineEdit_12.text()]
        wb = load_workbook("store.xlsx")
        name = self.comboBox_3.currentText()
        ws2 = wb[name]
        l1 = ws2.max_column
        for i in range(2,l1):
            e=i-2
            ws2.cell(row=row_modify,column=i).value= m_data[e]
            wb.save("store.xlsx")
        item = self.lineEdit_4.text()
        wb = load_workbook("store.xlsx")
        name = self.comboBox_3.currentText()
        ws2 = wb[name]
        ws3 = wb["Sheet"]
        o = 0
        d = 0
        prim_qunt = 0
        tota = 0
        for index, row in enumerate(ws3.iter_rows(min_row=2, values_only=True), start=2):
            val = row[2]
            if val == item:
                q = row[5]
                prim_qunt = int(q)
                ind = 1
                o = index
                if ind == 0:
                    QMessageBox.warning(self, "Erorr", "هذا الصنف غير موجود اضف الصنف اولا ")
                wb.save("store.xlsx")
        for index, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            row1 = list(row)
            val1 = row1[2]
            q1 = row1[5]
            if val1 == item:
                ty = row1[6]
                t =int(q1)
                d = index
                if ty == "اضافة":
                    tota = tota + t
                else:
                    tota = tota - t
                row1[10] = (tota + prim_qunt)
                ws2.cell(row=d, column=11).value = prim_qunt + tota
                wb.save("store.xlsx")
        total = prim_qunt + tota
        self.lineEdit_7.setText(str(total))
        ws3.cell(row=o, column=7).value = total
        #ws2.cell(row=d, column=11).value = total
        wb.save("store.xlsx")"""
class store_window (QMainWindow):
    def __init__(self):
        super(store_window,self).__init__()
        loadUi("stor2.ui",self)
        self.work()
        self.buttons()
    def update_combo(self):
        con = sqlite3.connect("store.db")
        cursor = con.cursor()
        items1 = []
        items2 = []
        items3 = []
        items4 = []
        items5 = []
        categ = self.comboBox_5.currentText()
        categ2 = self.comboBox_3.currentText()
        categ3 = self.comboBox_10.currentText()
        autdata=cursor.execute(f"SELECT type ,item_name FROM msheet")
        for  type, item_name in (autdata):
            if type == categ:
                items1.append(item_name)
        completer = QCompleter(items1)
        completer.setCaseSensitivity(False)
        self.lineEdit_7.setCompleter(completer)
        for  type, item_name in (autdata):
            if type == categ2:
                items2.append(item_name)
        completer2 = QCompleter(items2)
        completer2.setCaseSensitivity(False)
        self.lineEdit_4.setCompleter(completer2)
        for type, item_name in (autdata):
            if type == categ3:
                items3.append(item_name)
        completer3 = QCompleter(items3)
        completer3.setCaseSensitivity(False)
        self.lineEdit.setCompleter(completer3)
        name1 = self.comboBox_5.currentText()
        name2 = self.comboBox_3.currentText()
        table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy",
                      "الكهرباء": "electricity",
                      "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                      "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
        table1 = table_name[name1]
        table2 = table_name[name2]
        autoname1 = cursor.execute(f"SELECT source FROM {table1} ")
        for source in autoname1:
            s = list(source)
            for i in range(len(s)):
                items4.append(s[i])
        completer4 = QCompleter(items4)
        completer4.setCaseSensitivity(False)
        self.lineEdit_17.setCompleter(completer4)
        autoname2 = cursor.execute(f"SELECT source FROM {table2} ")
        for source in autoname2:
            s = list(source)
            for i in range(len(s)):
                items5.append(s[i])
        completer5 = QCompleter(items5)
        completer5.setCaseSensitivity(False)
        self.lineEdit_12.setCompleter(completer5)
        con.commit()
        con.close()
    def work(self):
        work_file="store.db"
        work_file1 = "store.xlsx"
        self.comboBox_5.currentTextChanged.connect(self.update_combo)
        self.comboBox_3.currentTextChanged.connect(self.update_combo)
        self.comboBox_10.currentTextChanged.connect(self.update_combo)
        if os.path.exists(work_file) and os.path.exists( work_file1):
            wb = load_workbook(work_file1)
            con=sqlite3.connect("store.db")
            cursor = con.cursor()
            items1 = []
            items2 = []
            items3 = []
            items4 = [" "]
            items5 = [" "]
            categ = self.comboBox_5.currentText()
            categ2 = self.comboBox_3.currentText()
            categ3 = self.comboBox_10.currentText()
            name1=self.lineEdit_17.text()
            autdata=cursor.execute(f"SELECT type ,item_name FROM msheet")
            for  type, item_name in autdata:
                if type == categ:
                    items1.append(item_name)
            completer = QCompleter(items1)
            completer.setCaseSensitivity(False)
            self.lineEdit_7.setCompleter(completer)
            autdata = cursor.execute(f"SELECT type ,item_name FROM msheet")
            for type, item_name in autdata:
                if type == categ2:
                    items2.append(item_name)
            completer2 = QCompleter(items2)
            completer2.setCaseSensitivity(False)
            self.lineEdit_4.setCompleter(completer2)
            autdata = cursor.execute(f"SELECT type ,item_name FROM msheet")
            for type, item_name in autdata:
                if type == categ3:
                    items3.append(item_name)
            completer3 = QCompleter(items3)
            completer3.setCaseSensitivity(False)
            self.lineEdit.setCompleter(completer3)
            name1 = self.comboBox_5.currentText()
            name2 = self.comboBox_3.currentText()
            table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy",
                          "الكهرباء": "electricity",
                          "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                          "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
            table1 = table_name[name1]
            table2 = table_name[name2]
            autoname1=cursor.execute(f"SELECT source FROM {table1} ")
            for source in autoname1:
                s=list(source)
                for i in range (len(s)):
                   items4.append(s[i])
            completer4 = QCompleter(items4)
            completer4.setCaseSensitivity(False)
            self.lineEdit_17.setCompleter(completer4)
            autoname2 = cursor.execute(f"SELECT source FROM {table2} ")
            for source in autoname2:
                s = list(source)
                for i in range(len(s)):
                    items5.append(s[i])
            completer5 = QCompleter(items5)
            completer5.setCaseSensitivity(False)
            self.lineEdit_12.setCompleter(completer5)
        else:
            con = sqlite3.connect("store.db")  # create database by connect
            # create tables
            cursor = con.cursor()  # connect chanel by data base
            cursor.execute(
                "CREATE TABLE msheet(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL ,balance REAL,number INTEGER,date DATE,source TEXT)")  # داله التنفيذ لاي امر
            cursor.execute(
                "CREATE TABLE painting(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE rwaked(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE sehy(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE electricity(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE petrol(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE gates(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE books(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE wood(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE finished(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            cursor.execute(
                "CREATE TABLE used(code INTEGER , type TEXT,item_name TEXT ,unit TEXT ,state TEXT , amount REAL,process TEXT,number INTEGER,date DATE,source TEXT,balance REAL)")
            con.commit()
            con.close()
            wb = Workbook()
            sheets = wb.sheetnames
            sheet = sheets[0]
            ws = wb[sheet]
            titles = ["الكود", "نوع الصنف", "اسم الصنف", "الوحدة", "حالة الصنف", "الكمية", "الرصيد", "رقم الاذن",
                      "تاريخ الاضافة", "المصدر"]
            ws.append(titles)
            types = ["دهانات", "مهمات راكدة", 'السباكة', 'الكهرباء', 'مواد بتروليه',
                     'البوابات والمعدات', 'دفاتر وجنين', 'النجارة', "الخردة", 'مستعملة']
            sheet_titles = ["الكود", 'نوع الصنف', 'اسم الصنف', 'الوحدة', 'حالة الصنف', 'الكمية', 'نوع الحركة',
                            'رقم الاذن', 'التاريخ', 'المصدر', 'الرصيد']
            for i in range(0, len(types)):
                wb.create_sheet(index=(i + 1), title=types[i])
            for i in range(1, len(types) + 1):
                wsh = wb.sheetnames[i]
                wb[wsh].append(sheet_titles)
            wb.save(work_file1)
    def change_style(self):
        self.tableWidget.setColumnWidth(0,20)
        self.tableWidget.setColumnWidth(0, 100)
        self.tableWidget.setColumnWidth(0, 50)
        self.tableWidget.setColumnWidth(0, 50)
        self.tableWidget.setColumnWidth(0, 50)
        self.tableWidget.setColumnWidth(0, 50)
    def buttons(self):
        self.pushButton.clicked.connect(self.cleartabes)
        self.add_new_item_btn.clicked.connect(self.add_new_item)
        self.pushButton_2.clicked.connect(self.add_quantitiy)
        self.pushButton_3.clicked.connect(self.cut_quantity)
        self.pushButton_5.clicked.connect(self.search)
        self.pushButton_4.clicked.connect(self.show_data)
        self.tableWidget.cellClicked.connect(self.on_cell_clicked)
        self.tableWidget_2.cellClicked.connect(self.on_cell_clicked1)
    def add_new_item(self):
        if self.lineEdit_2.text() =="" or self.lineEdit_3.text()=="" or self.lineEdit_14.text()==""or self.lineEdit_15.text()=="":
            QMessageBox.about(self, "Erorr", "من فضلك املاء البيانات ")
        else:
            con = sqlite3.connect("store.db")
            cursor = con.cursor()
            a=0
            rowd=cursor.execute("SELECT rowid,* FROM msheet")
            for row in rowd:
                a=row[0] +1
            new_item = [a, self.comboBox.currentText(), self.lineEdit_2.text(), self.comboBox_2.currentText(),
                        self.comboBox_7.currentText(),
                        float(self.lineEdit_3.text()), float(self.lineEdit_3.text()),self.lineEdit_14.text(),self.dateEdit_2.text(),self.lineEdit_15.text()]
            cursor.execute("INSERT INTO msheet VALUES(?,?,?,?,?,?,?,?,?,?)",new_item)
            con.commit()
            con.close()
            wb = load_workbook("store.xlsx")
            sheets = wb.sheetnames
            sheet = sheets[0]
            ws1 = wb[sheet]
            ro_num = ws1.max_row + 1
            new_item = [ro_num, self.comboBox.currentText(), self.lineEdit_2.text(), self.comboBox_2.currentText(),
                        self.comboBox_7.currentText(),
                        float(self.lineEdit_3.text()), float(self.lineEdit_3.text()),self.lineEdit_14.text(),self.dateEdit_2.text(),self.lineEdit_15.text()]
            ws1.append(new_item)
            wb.save("store.xlsx")
            self.lineEdit_2.setText("")
            self.lineEdit_3.setText("")
            self.lineEdit_14.setText("")
            self.lineEdit_15.setText("")
    def add_quantitiy(self):
        if self.lineEdit_4.text() =="" or self.lineEdit_5.text()=="" or self.lineEdit_12.text()=="" or self.lineEdit_10.text()=="":
            QMessageBox.about(self, "Erorr", "من فضلك املاء البيانات ")
            return
        else:
            table_name={"دهانات":"painting","مهمات راكدة":"rwaked","السباكة":"sehy","الكهرباء":"electricity","مواد بتروليه":"petrol","البوبات والمعدات":"gates","دفاتر وجناين":"books","النجارة":"wood","الخردة":"finished","مستعمله":"used"}
            item=self.lineEdit_4.text()
            prim_qunt=0
            tota= 0
            total=0
            #add=[self.comboBox_3.currentText(),self.lineEdit_4.text(),self.comboBox_4.currentText(),self.comboBox_8.currentText(),int(self.lineEdit_5.text()),"اضافة",self.lineEdit_10.text(),self.dateEdit.text(),self.lineEdit_12.text()]
            name = self.comboBox_3.currentText()
            table=table_name[name]
            con = sqlite3.connect("store.db")
            cursor = con.cursor()
            prim_unit=0
            ind = 0
            cursor.execute("SELECT rowid, amount FROM msheet WHERE item_name=?", (item,))
            row = cursor.fetchone()
            if row is None:
                QMessageBox.warning(self, "Erorr", "هذا الصنف غير موجود اضف الصنف اولا ")
                return
            else:
                rid, prim_unit = row
                prim_unit = float(prim_unit)
                cursor.execute(f"SELECT MAX(rowid) FROM {table}")
                last_id = cursor.fetchone()[0]
                if last_id is None:
                    last_id = 0
                add = [last_id+1, self.comboBox_3.currentText(), self.lineEdit_4.text(), self.comboBox_4.currentText(),
                      self.comboBox_8.currentText(), float(self.lineEdit_5.text()), "اضافة", self.lineEdit_10.text(),
                      self.dateEdit.text(),self.lineEdit_12.text(), 0]
                cursor.execute(f"INSERT INTO {table}(code, type,item_name,unit,state, amount,process,number,date,source,balance) VALUES(?,?,?,?,?,?,?,?,?,?,?)", add)
               # con.commit()
                cursor.execute(f"SELECT rowid, amount,process FROM {table} WHERE item_name=?", (item,))
                rows =cursor.fetchall()
                for rowid, amount, process in rows:
                    if process == "اضافة":
                        tota = tota + float(amount)
                    else:
                        tota = tota - float(amount)
                total =float (tota + prim_unit)
                cursor.execute(f"UPDATE {table} SET balance =? WHERE rowid=?",(total,rowid,))
                cursor.execute(f"UPDATE msheet SET balance =?  WHERE rowid= ?",(total,rid,))
                con.commit()
                self.lineEdit_6.setText(str(total))
                wb = load_workbook("store.xlsx")
                name=self.comboBox_3.currentText()
                ws2 = wb[name]
                next_row = ws2.max_row
                add = [next_row,self.comboBox_3.currentText(), self.lineEdit_4.text(), self.comboBox_4.currentText(),
                       self.comboBox_8.currentText(), float(self.lineEdit_5.text()), "اضافة", self.lineEdit_10.text(),
                       self.dateEdit.text(), self.lineEdit_12.text(),total]
                ws2.append(add)
                ws3=wb["Sheet"]
                for row in ws3.iter_rows(min_row=2, values_only=False):
                    if row[2].value == item:
                        row[6].value = total  # العمود السابع = الرصيد
                        break
                wb.save("store.xlsx")
                con.commit()
                con.close()

                self.lineEdit_4.clear()
                self.lineEdit_5.clear()
                self.lineEdit_10.clear()
                self.lineEdit_12.clear()
    def cut_quantity(self):
        if self.lineEdit_7.text() =="" or self.lineEdit_8.text()=="" or self.lineEdit_18.text()=="" or self.lineEdit_17.text()=="":
            QMessageBox.about(self, "Erorr", "من فضلك املاء البيانات ")
            return

        table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy", "الكهرباء": "electricity",
                          "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                          "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
        item = self.lineEdit_7.text()
        prim_qunt = 0
        tota = 0
        name = self.comboBox_5.currentText()
        table = table_name[name]
        con = sqlite3.connect("store.db")
        cursor = con.cursor()
        prim_unit = 0
        ind = 0
        cursor.execute("SELECT rowid, amount FROM msheet WHERE item_name=?", (item,))
        row = cursor.fetchone()
        if row is None:
            QMessageBox.warning(self, "Erorr", "هذا الصنف غير موجود اضف الصنف اولا ")
            return

        rid, prim_unit = row
        prim_unit = float(prim_unit)
        cursor.execute(f"SELECT MAX(rowid) FROM {table}")
        last_id = cursor.fetchone()[0]
        if last_id is None:
            last_id = 0
        add = [last_id + 1, self.comboBox_5.currentText(), self.lineEdit_7.text(), self.comboBox_6.currentText(),
               self.comboBox_12.currentText(), float(self.lineEdit_8.text()), "صرف", self.lineEdit_18.text(),
               self.dateEdit_3.text(), self.lineEdit_17.text(), 0]
        cursor.execute(
            f"INSERT INTO {table}(code, type,item_name,unit,state, amount,process,number,date,source,balance) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            add)
        cursor.execute(f"SELECT rowid, amount,process FROM {table} WHERE item_name=?", (item,))
        rows = cursor.fetchall()
        for rowid, amount, process in rows:
            if process == "اضافة":
                tota = tota + float(amount)
            else:
                tota = tota - float(amount)
        total = float(tota + prim_unit)
        #rowd2 = cursor.execute(f"SELECT rowid, * FROM {table} ORDER BY rowid DESC LIMIT 1")
        if total < 0:
            # حذف آخر عملية صرف (لأنها غير مسموحة)
            cursor.execute(f"DELETE FROM {table} WHERE rowid=(SELECT MAX(rowid) FROM {table})")
            QMessageBox.warning(self, "Erorr",
                                f"الرصيد الحالي هو {prim_unit + (tota + float(self.lineEdit_8.text()))} لا يسمح بصرف هذه الكمية")
        else:
            cursor.execute(f"UPDATE {table} SET balance = ? WHERE rowid=(SELECT MAX(rowid) FROM {table})", (total,))
            cursor.execute("UPDATE msheet SET balance = ? WHERE rowid=?", (total, rid))
            self.lineEdit_9.setText(str(total))

            # تحديث في ملف Excel
            wb = load_workbook("store.xlsx")
            ws2 = wb[name]
            ws3 = wb["Sheet"]

            # إضافة الصف في شيت النوع
            next_row = ws2.max_row
            add_xl = [
                next_row, name, item, self.comboBox_6.currentText(),
                self.comboBox_12.currentText(), float(self.lineEdit_8.text()),
                "صرف", self.lineEdit_18.text(), self.dateEdit_3.text(),
                self.lineEdit_17.text(), total
            ]
            ws2.append(add_xl)
            # تحديث الرصيد في الشيت الأساسي
            for row in ws3.iter_rows(min_row=2, values_only=False):
                if row[2].value == item:
                    row[6].value = total  # العمود السابع = الرصيد
                    break
            wb.save("store.xlsx")
            con.commit()
            con.close()
            self.lineEdit_7.clear()
            self.lineEdit_8.clear()
            self.lineEdit_18.clear()
            self.lineEdit_17.clear()
    def cleartabes(self):
        self.tableWidget_2.clearContents()
        self.tableWidget_2.verticalHeader().setVisible(False)
        self.lineEdit.setText("")
        self.lineEdit_11.setText("")
        self.lineEdit_13.setText("")
    def search(self):
        self.tableWidget_2.clearContents()
        self.tableWidget_2.verticalHeader().setVisible(False)
        self.lineEdit_11.setText("")
        self.lineEdit_13.setText("")
        item_name=self.lineEdit.text()
        name = self.comboBox_10.currentText()
        table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy", "الكهرباء": "electricity",
                      "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                      "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
        table = table_name[name]
        con = sqlite3.connect("store.db")
        cursor = con.cursor()
        rows = cursor.execute(f"SELECT code,item_name, amount,process,number,date,source,balance FROM {table} WHERE item_name=?", (item_name,))
        rn = 0
        self.row1=0
        for r5 in rows:
            self.tableWidget_2.setRowCount(rn+1)
            row2=list(r5)
            for i in range(len(row2)) :
                self.tableWidget_2.setItem(self.row1,i, QTableWidgetItem(str(row2[i])))
            self.row1 += 1
            rn += 1
        #self.tableWidget_2.resizeColumnsToContents()
        """wb = load_workbook("store.xlsx")
        name = self.comboBox_10.currentText()
        ws2 = wb[name]
        r2 = ws2.max_row + 1
        l2 = ws2.max_column + 1
        #self.tableWidget_2.setRowCount(r2)
        self.row = 0
        ws3 = wb["Sheet"]
        r3 = ws3.max_row + 1
        rn = 1
        y=0
        for i1 in range(2, r3):
            val11 = ws3.cell(row=i1, column=3).value
            if val11 == item_name:
                a1=i1
                self.prim_c=ws3.cell(row=a1, column=6).value
                self.raseed=ws3.cell(row=a1, column=7).value"""
        cursor.execute(f"SELECT amount FROM msheet  WHERE item_name=? ORDER BY rowid DESC LIMIT 1", (item_name,))
        r=cursor.fetchone()
        for amount in r:
            self.lineEdit_11.setText(str(amount))
        cursor.execute(f"SELECT balance FROM {table} WHERE item_name=? ORDER BY rowid DESC LIMIT 1", (item_name,))
        r2 = cursor.fetchone()
        if r2 is not None :
            for balance in r2:
                self.lineEdit_13.setText(str(balance))
        else:
            self.lineEdit_13.setText(str(amount))
        for row in range(self.tableWidget_2.rowCount()):
          for col in range(self.tableWidget_2.columnCount()):
            item = self.tableWidget_2.item(row, col)
            if item is None:
                item = QTableWidgetItem("")
                self.tableWidget_2.setItem(row, col, item)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
    def show_message(self):
            msg=QMessageBox()
            msg.setWindowTitle("Question")
            msg.setText("اختر العمليه المطلوبه ")
            msg.setIcon(QMessageBox.Question)
            showbtn=msg.addButton("عرض البيانات",QMessageBox.ActionRole)
            delbtn = msg.addButton("حذف",QMessageBox.ActionRole)
            msg.addButton(QMessageBox.Cancel)
            msg.exec_()
            if msg.clickedButton()==showbtn:
                print("show")
            elif msg.clickedButton()==delbtn:
                print("del")
    def show_data(self):
        self.tableWidget.clearContents()
        self.tableWidget.verticalHeader().setVisible(False)
        self.row1 = 0
        name = self.comboBox_9.currentText()
        table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy", "الكهرباء": "electricity",
                      "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                      "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
        table = table_name[name]
        con = sqlite3.connect("store.db")
        cursor = con.cursor()
        rows=cursor.execute(f"SELECT * FROM msheet WHERE type=?",(name,))
        rn=1
        for r in rows:
            self.tableWidget.setRowCount(rn)
            row1=list(r)
            for i in range (len(row1)) :
                self.tableWidget.setItem(self.row1,i, QTableWidgetItem(str(row1[i])))
            self.row1 += 1
            rn += 1
        """ wb = load_workbook("store.xlsx")
        ws2 = wb["Sheet"]
        r2 = ws2.max_row + 1
        l2 = ws2.max_column + 1
        #self.tableWidget.setRowCount(r2)"""
        #rn=1
        """for index, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            if row[1]==item_cat:
                self.tableWidget.setRowCount(rn)
                rn+=1
                f=index
                da=[]
                for v in range (1,l2):
                    va=ws2.cell(row=f, column=v).value
                    da.append(va)
                for e in range(0, len(da)):
                    #self.tableWidget.setRowHight(1, 20)
                    self.tableWidget.setItem(self.row1, e, QTableWidgetItem(str(da[e])))
                self.row1 += 1"""
        for row in range(self.tableWidget.rowCount()):
          for col in range(self.tableWidget.columnCount()):
            item = self.tableWidget.item(row, col)
            if item is None:
                item = QTableWidgetItem("")
                self.tableWidget.setItem(row, col, item)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
    def on_cell_clicked(self,row, column):
            valu = self.tableWidget.item(row,2).text()
            msg = QMessageBox()
            msg.setWindowTitle("Question")
            msg.setText("اختر العمليه المطلوبه ")
            msg.setIcon(QMessageBox.Question)
            showbtn = msg.addButton("عرض البيانات", QMessageBox.ActionRole)
            delbtn = msg.addButton("حذف", QMessageBox.ActionRole)
            msg.addButton(QMessageBox.Cancel)
            msg.exec_()
            if msg.clickedButton() == showbtn:
                sheet=self.comboBox_9.currentText()
                self.comboBox_10.setCurrentText(sheet)
                self.lineEdit.setText(valu)
                self.search()
                self.tabWidget.setCurrentIndex(2)
            elif msg.clickedButton() == delbtn:
                name = self.comboBox_9.currentText()
                table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy",
                              "الكهرباء": "electricity",
                              "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                              "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
                table = table_name[name]
                con = sqlite3.connect("store.db")
                cursor = con.cursor()
                cursor.execute(f"DELETE FROM msheet WHERE item_name=?", (valu,))
                cursor.execute(f"DELETE FROM {table} WHERE item_name=?", (valu,))
                wb = load_workbook("store.xlsx")
                ws2 = wb["Sheet"]
                r2 = ws2.max_row + 1
                ws = wb[name]
                for i in range(2, ws2.max_row + 1):
                    for j in range(1, ws2.max_column + 1):
                        ws2.cell(row=i, column=j).value = ""
                rin2 = 2
                dataa = cursor.execute("SELECT  *  FROM msheet ")
                for row in dataa:
                    data2 = list(row)
                    for j in range(0, len(data2)):
                        ws2.cell(row=rin2, column=j + 1).value = data2[j]
                    rin2 += 1
                for i in range(2, ws.max_row + 1):
                    for j in range(1, ws.max_column + 1):
                        ws.cell(row=i, column=j).value = ""
                rin = 2
                data = cursor.execute(f"SELECT  *  FROM {table} ")
                for row in data:
                    data1 = list(row)
                    for j in range(0, len(data1)):
                        ws.cell(row=rin, column=j + 1).value = data1[j]
                    rin += 1
                wb.save("store.xlsx")
                con.commit()
                con.close()
    def on_cell_clicked1(self, row, column):
        global row_modify
        valu = self.tableWidget_2.item(row, 0).text()
        item=self.tableWidget_2.item(row, 1).text()
        pros=self.tableWidget_2.item(row, 3).text()
        row_modify=int(valu)+1
        msg = QMessageBox()
        msg.setWindowTitle("Question")
        msg.setText("اختر العمليه المطلوبه ")
        msg.setIcon(QMessageBox.Question)
        modbtn = msg.addButton("تعديل البيانات", QMessageBox.ActionRole)
        delbtn = msg.addButton("حذف العملية", QMessageBox.ActionRole)
        msg.addButton(QMessageBox.Cancel)
        msg.exec_()
        if msg.clickedButton() == delbtn:
            name = self.comboBox_10.currentText()
            table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy",
                          "الكهرباء": "electricity",
                          "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                          "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
            table = table_name[name]
            con = sqlite3.connect("store.db")
            cursor = con.cursor()
            cursor.execute(f"DELETE FROM {table} WHERE code=?", (valu,))
            cursor.execute(f"SELECT rowid, amount,process ,balance FROM {table} WHERE item_name=?", (item,))
            data =cursor.fetchall()
            prim_unit =float(self.lineEdit_11.text())
            tota=0
            total=0
            for rowid, amount , process, balance in data:
                if process == "اضافة":
                    tota = tota + float(amount)
                else:
                    tota = tota - float(amount)
                total = float(tota + prim_unit)
                cursor.execute(f"UPDATE {table} SET balance =? WHERE rowid=?", (total, rowid,))
            cursor.execute(f"UPDATE msheet SET balance =? WHERE item_name=?", (total, item,))
            rows = cursor.execute(f"SELECT rowid FROM {table} ORDER BY rowid").fetchall()
            i = 1
            for (rowid,) in rows:
                cursor.execute(f"UPDATE {table} SET code=? WHERE rowid=?", (i, rowid))
                i += 1
            cursor.execute(f"SELECT COUNT(*) FROM {table}")
            row_count = cursor.fetchone()[0]
            wb = load_workbook("store.xlsx")
            ws = wb[name]
            for i in range(2, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=i, column=j).value = ""
            r = row_count
            rin = 2
            data = cursor.execute(f"SELECT  *  FROM {table} ")
            for row in data:
                data1 = list(row)
                for j in range(0, len(data1)):
                    ws.cell(row=rin, column=j + 1).value = data1[j]
                rin += 1
            wb.save("store.xlsx")
            con.commit()
            con.close()
        elif  msg.clickedButton() == modbtn:
            global cod
            global name1
            state=self.tableWidget_2.item(row, 3).text()
            cod=self.tableWidget_2.item(row, 0).text()
            name1 = self.comboBox_10.currentText()
            table_name = {"دهانات": "painting", "مهمات راكدة": "rwaked", "السباكة": "sehy",
                          "الكهرباء": "electricity",
                          "مواد بتروليه": "petrol", "البوبات والمعدات": "gates", "دفاتر وجناين": "books",
                          "النجارة": "wood", "الخردة": "finished", "مستعمله": "used"}
            table = table_name[name1]
            con = sqlite3.connect("store.db")
            cursor = con.cursor()
            con.commit()
            con.close()
            self.modify_window1 = modify_window()
            self.modify_window1.show()
            #self.lineEdit_4.setText=data[1]
if __name__=="__main__":
    app=QApplication(sys.argv)
    window=store_window()
    window.show()
    sys.exit(app.exec())